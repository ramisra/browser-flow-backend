"""Unit tests for Excel tools."""

import pytest
from pathlib import Path

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.core.tools.excel_tools import ExcelTools


@pytest.mark.asyncio
async def test_create_excel_file_basic(excel_tools, sample_extraction_data):
    """Test basic Excel file creation."""
    columns = ["name", "price", "stock"]
    file_path = await excel_tools.create_excel_file(
        data=sample_extraction_data,
        columns=columns,
    )

    # Validations
    assert Path(file_path).exists()
    assert Path(file_path).suffix == ".xlsx"

    # Read back and validate content
    if PANDAS_AVAILABLE:
        df = pd.read_excel(file_path, engine="openpyxl")
    elif OPENPYXL_AVAILABLE:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        df_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            df_data.append(dict(zip(headers, row)))
        # Create simple dict-based validation
        assert len(df_data) == 2
        assert df_data[0]["name"] == "Product A"
        assert df_data[0]["price"] == 100
        assert df_data[1]["name"] == "Product B"
        return
    
    df = pd.read_excel(file_path, engine="openpyxl")
    assert len(df) == 2
    assert list(df.columns) == columns
    assert df.iloc[0]["name"] == "Product A"
    assert df.iloc[0]["price"] == 100
    assert df.iloc[1]["name"] == "Product B"


@pytest.mark.asyncio
async def test_create_excel_with_inferred_columns(excel_tools, sample_extraction_data):
    """Test Excel creation with inferred columns."""
    file_path = await excel_tools.create_excel_file(
        data=sample_extraction_data,
        columns=None,  # Should infer from data
    )

    assert Path(file_path).exists()
    df = pd.read_excel(file_path, engine="openpyxl")
    assert len(df) == 2
    assert "name" in df.columns
    assert "price" in df.columns
    assert "stock" in df.columns


@pytest.mark.asyncio
async def test_create_excel_with_special_characters(excel_tools):
    """Test Excel creation with special characters and unicode."""
    data = [
        {"name": "José", "price": 100, "description": "Product with émojis 🎉"},
        {"name": "Müller", "price": 200, "description": "Special chars: <>&"},
    ]
    file_path = await excel_tools.create_excel_file(data=data)

    assert Path(file_path).exists()
    df = pd.read_excel(file_path, engine="openpyxl")
    assert len(df) == 2
    assert "José" in df["name"].values
    assert "Müller" in df["name"].values


@pytest.mark.asyncio
async def test_append_to_excel(excel_tools, sample_extraction_data):
    """Test appending data to existing Excel file."""
    # Create initial file
    columns = ["name", "price", "stock"]
    file_path = await excel_tools.create_excel_file(
        data=sample_extraction_data,
        columns=columns,
    )

    # Append new data
    new_data = [
        {"name": "Product C", "price": 300, "stock": 20},
    ]
    await excel_tools.append_to_excel(file_path, new_data, columns)

    # Validate
    df = pd.read_excel(file_path, engine="openpyxl")
    assert len(df) == 3
    assert df.iloc[2]["name"] == "Product C"


@pytest.mark.asyncio
async def test_read_excel(excel_tools, sample_extraction_data):
    """Test reading data from Excel file."""
    file_path = await excel_tools.create_excel_file(
        data=sample_extraction_data,
        columns=["name", "price", "stock"],
    )

    data = await excel_tools.read_excel(file_path)
    assert len(data) == 2
    assert data[0]["name"] == "Product A"
    assert data[1]["price"] == 200


@pytest.mark.asyncio
async def test_create_excel_empty_data_raises_error(excel_tools):
    """Test that creating Excel with empty data raises error."""
    with pytest.raises(ValueError, match="Data list cannot be empty"):
        await excel_tools.create_excel_file(data=[])


@pytest.mark.asyncio
async def test_create_excel_custom_filename(excel_tools, sample_extraction_data):
    """Test creating Excel with custom filename."""
    file_path = await excel_tools.create_excel_file(
        data=sample_extraction_data,
        file_name="custom_test.xlsx",
    )

    assert Path(file_path).name == "custom_test.xlsx"
    assert Path(file_path).exists()
