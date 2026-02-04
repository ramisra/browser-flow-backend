"""Excel writing tools for data extraction."""

import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    try:
        import pandas as pd
        PANDAS_AVAILABLE = True
    except ImportError:
        PANDAS_AVAILABLE = False


class ExcelTools:
    """Tools for creating and managing Excel files."""

    def __init__(self, storage_dir: Optional[str] = None):
        """Initialize Excel tools.

        Args:
            storage_dir: Directory to store Excel files (default: app/storage/excel)
        """
        if storage_dir:
            self.storage_dir = Path(storage_dir)
        else:
            self.storage_dir = Path("app/storage/excel")

        # Create storage directory if it doesn't exist
        self.storage_dir.mkdir(parents=True, exist_ok=True)

        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            raise ImportError(
                "Neither openpyxl nor pandas is available. "
                "Please install one of them: pip install openpyxl or pip install pandas"
            )

    async def create_excel_file(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        file_name: Optional[str] = None,
        sheet_name: Optional[str] = None,
    ) -> str:
        """Create an Excel file with data.

        Args:
            data: List of dictionaries with data
            columns: Optional list of column names (inferred from data if not provided)
            file_name: Optional file name (generated if not provided)
            sheet_name: Optional sheet name (default: Data)

        Returns:
            Path to created Excel file

        Raises:
            ValueError: If data is empty or invalid
        """
        if not data:
            raise ValueError("Data list cannot be empty")

        # Infer columns from data if not provided
        if not columns:
            columns = list(data[0].keys()) if data else []

        # Generate file name if not provided
        if not file_name:
            file_name = f"data_{uuid.uuid4().hex[:8]}.xlsx"

        file_path = self.storage_dir / file_name

        # Use openpyxl if available, otherwise pandas
        if OPENPYXL_AVAILABLE:
            await self._create_with_openpyxl(
                data, columns, file_path, sheet_name
            )
        elif PANDAS_AVAILABLE:
            await self._create_with_pandas(
                data, columns, file_path, sheet_name
            )

        return str(file_path)

    async def _create_with_openpyxl(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        file_path: Path,
        sheet_name: Optional[str] = None,
    ) -> None:
        """Create Excel file using openpyxl."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name or "Data"

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)

    async def _create_with_pandas(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        file_path: Path,
        sheet_name: Optional[str] = None,
    ) -> None:
        """Create Excel file using pandas."""
        # Ensure all rows have the same columns
        normalized_data = []
        for row in data:
            normalized_row = {col: row.get(col, "") for col in columns}
            normalized_data.append(normalized_row)

        df = pd.DataFrame(normalized_data, columns=columns)
        df.to_excel(
            file_path,
            index=False,
            engine="openpyxl",
            sheet_name=sheet_name or "Data",
        )

    async def append_to_excel(
        self,
        file_path: str,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        sheet_name: Optional[str] = None,
    ) -> None:
        """Append data to an existing Excel file.

        Args:
            file_path: Path to Excel file
            data: List of dictionaries with data to append
            columns: Optional list of column names
            sheet_name: Optional sheet name (default: Data)

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If data is empty
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if not data:
            raise ValueError("Data list cannot be empty")

        if not columns:
            columns = list(data[0].keys()) if data else []

        if OPENPYXL_AVAILABLE:
            await self._append_with_openpyxl(
                path, data, columns, sheet_name
            )
        elif PANDAS_AVAILABLE:
            await self._append_with_pandas(
                path, data, columns, sheet_name
            )

    async def _append_with_openpyxl(
        self,
        file_path: Path,
        data: List[Dict[str, Any]],
        columns: List[str],
        sheet_name: Optional[str] = None,
    ) -> None:
        """Append data using openpyxl."""
        wb = load_workbook(file_path)
        target_sheet = sheet_name or "Data"
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
            header_row = 1
            next_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=target_sheet)
            header_row = 1
            next_row = 2
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=header_row, column=col_idx, value=col_name)

        # Write data
        for row_data in data:
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                ws.cell(row=next_row, column=col_idx, value=value)
            next_row += 1

        wb.save(file_path)

    async def _append_with_pandas(
        self,
        file_path: Path,
        data: List[Dict[str, Any]],
        columns: List[str],
        sheet_name: Optional[str] = None,
    ) -> None:
        """Append data using pandas."""
        # Read existing data
        existing_df = pd.read_excel(
            file_path, engine="openpyxl", sheet_name=sheet_name or 0
        )

        # Normalize new data
        normalized_data = []
        for row in data:
            normalized_row = {col: row.get(col, "") for col in columns}
            normalized_data.append(normalized_row)

        # Append
        new_df = pd.DataFrame(normalized_data, columns=columns)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df.to_excel(
            file_path,
            index=False,
            engine="openpyxl",
            sheet_name=sheet_name or "Data",
        )

    async def read_excel(
        self, file_path: str
    ) -> List[Dict[str, Any]]:
        """Read data from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of dictionaries with data

        Raises:
            FileNotFoundError: If file doesn't exist
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if PANDAS_AVAILABLE:
            df = pd.read_excel(file_path, engine="openpyxl")
            return df.to_dict("records")
        elif OPENPYXL_AVAILABLE:
            wb = load_workbook(file_path)
            ws = wb.active

            # Read headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Read data
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                data.append(row_dict)

            return data
        else:
            raise ImportError("Neither openpyxl nor pandas is available")


__all__ = ["ExcelTools"]
